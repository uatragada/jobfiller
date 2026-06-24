from __future__ import annotations

import json
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Job
from ..settings import OUTPUT_ROOT
from .processor import latest_artifact, latest_grade
from .time_utils import scan_sort_key


WORKBOOK_PATH = OUTPUT_ROOT / "jobfiller-feedback-loop.xlsx"
JSON_EXPORT_PATH = OUTPUT_ROOT / "jobfiller-feedback-loop.json"
CSV_EXPORT_PATH = OUTPUT_ROOT / "jobfiller-feedback-loop.csv"


def _add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "Company": 18,
        "Title": 36,
        "Location": 18,
        "Key Requirements": 42,
        "Keywords": 36,
        "Salary": 18,
        "Materials Needed": 40,
        "Source URL": 52,
        "Apply URL": 52,
        "Resume PDF Path": 58,
        "Resume LaTeX Path": 58,
        "Cover Letter Path": 58,
        "Manual Writing Questions": 42,
        "Local LLM Risks": 70,
        "Cover Letter Text": 90,
        "Action": 56,
        "Files": 60,
    }
    for idx, column_cells in enumerate(sheet.columns, start=1):
        header = str(column_cells[0].value)
        sheet.column_dimensions[get_column_letter(idx)].width = widths.get(header, 18)
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"


def export_current_workbook(db: Session, path: Path = WORKBOOK_PATH) -> Path:
    jobs = ordered_jobs(db)
    workbook = Workbook()
    workbook.remove(workbook.active)

    job_rows = []
    tailoring_rows = []
    cover_rows = []
    checklist_rows = []

    for index, job in enumerate(jobs, start=1):
        artifact = latest_artifact(job)
        grade = latest_grade(job)
        open_questions = [q.question_text for q in job.questions if q.status == "OPEN"]
        manual_questions = list(dict.fromkeys([q.question_text for q in job.questions if q.status == "OPEN"]))
        if job.manual_questions:
            manual_questions.insert(0, job.manual_questions)
        risks = []
        if grade and grade.risks_json:
            risks = json.loads(grade.risks_json)
        resume_pdf = artifact.resume_pdf_path if artifact else ""
        resume_tex = artifact.resume_tex_path if artifact else ""
        cover_path = artifact.cover_letter_path if artifact else ""
        cover_text = Path(cover_path).read_text(encoding="utf-8") if cover_path and Path(cover_path).exists() else ""
        compile_status = artifact.compile_status if artifact else ""

        job_rows.append(
            [
                index,
                job.company,
                job.title,
                job.location,
                job.work_model,
                job.source_url,
                job.apply_url,
                job.status,
                job.fit_score,
                grade.overall_grade if grade else "",
                "yes" if grade and grade.ready_to_send else "no",
                job.key_requirements,
                job.keywords,
                job.salary,
                job.materials,
                resume_pdf,
                resume_tex,
                cover_path,
                "\n".join(manual_questions) or "Resume, cover letter, portfolio, or other materials if requested",
                "\n".join(open_questions),
                job.posting_age_text,
                job.posted_at.isoformat() if job.posted_at else "",
                compile_status,
                "\n".join(risks),
            ]
        )
        tailoring_rows.append(
            [
                index,
                job.company,
                job.title,
                job.role_family,
                job.key_requirements,
                job.keywords,
                compile_status,
                resume_pdf,
            ]
        )
        cover_rows.append([index, job.company, job.title, cover_text, cover_path])
        checklist_rows.append(
            [
                index,
                job.company,
                job.title,
                "Open apply URL, confirm the role is still accepting applications, upload the tailored resume, and paste or attach the cover letter if requested.",
                f"{resume_pdf}\n{cover_path}",
                "\n".join(open_questions) or "Review all required form fields manually, including work authorization and sponsorship.",
            ]
        )

    _add_sheet(
        workbook,
        "Jobs",
        [
            "Apply Order",
            "Company",
            "Title",
            "Location",
            "Work Model",
            "Source URL",
            "Apply URL",
            "Status",
            "Fit Score",
            "Local LLM Grade",
            "Ready To Send",
            "Key Requirements",
            "Keywords",
            "Salary",
            "Materials Needed",
            "Resume PDF Path",
            "Resume LaTeX Path",
            "Cover Letter Path",
            "Manual Writing Questions",
            "Posting Age",
            "Posted At",
            "Compile Status",
            "Local LLM Risks",
        ],
        job_rows,
    )
    _add_sheet(
        workbook,
        "Resume Tailoring",
        ["Apply Order", "Company", "Title", "Role Family", "Target Requirements", "Target Keywords", "Compile Status", "Resume PDF Path"],
        tailoring_rows,
    )
    _add_sheet(workbook, "Cover Letters", ["Apply Order", "Company", "Title", "Cover Letter Text", "Cover Letter Path"], cover_rows)
    _add_sheet(workbook, "Tomorrow Checklist", ["Apply Order", "Company", "Title", "Action", "Files", "Manual Checks"], checklist_rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def current_export_rows(db: Session) -> list[dict[str, object]]:
    jobs = ordered_jobs(db)
    rows: list[dict[str, object]] = []
    for job in jobs:
        artifact = latest_artifact(job)
        grade = latest_grade(job)
        rows.append(
            {
                "id": job.id,
                "company": job.company,
                "title": job.title,
                "location": job.location,
                "work_model": job.work_model,
                "source": job.source,
                "source_url": job.source_url,
                "apply_url": job.apply_url,
                "status": job.status,
                "fit_score": job.fit_score,
                "grade": grade.overall_grade if grade else "",
                "ready_to_send": grade.ready_to_send if grade else False,
                "key_requirements": job.key_requirements,
                "keywords": job.keywords,
                "salary": job.salary,
                "materials": job.materials,
                "manual_questions": job.manual_questions,
                "posting_age_text": job.posting_age_text,
                "posted_at": job.posted_at.isoformat() if job.posted_at else "",
                "last_seen_at": job.last_seen_at.isoformat() if job.last_seen_at else "",
                "resume_pdf_path": artifact.resume_pdf_path if artifact else "",
                "resume_tex_path": artifact.resume_tex_path if artifact else "",
                "cover_letter_path": artifact.cover_letter_path if artifact else "",
                "open_questions": sum(1 for question in job.questions if question.status == "OPEN"),
                "notes": job.notes,
            }
        )
    return rows


def ordered_jobs(db: Session) -> list[Job]:
    jobs = db.scalars(select(Job)).all()
    return sorted(
        jobs,
        key=lambda job: scan_sort_key(
            job.posted_at,
            job.first_seen_at,
            job.fit_score,
            remote_first=True,
            location=job.location,
            work_model=job.work_model,
        ),
        reverse=True,
    )


def build_tomorrow_checklist(db: Session) -> list[dict[str, object]]:
    checklist: list[dict[str, object]] = []
    for index, job in enumerate(ordered_jobs(db), start=1):
        artifact = latest_artifact(job)
        if not artifact:
            continue
        grade = latest_grade(job)
        open_questions = [q.question_text for q in job.questions if q.status == "OPEN"]
        manual_questions = list(dict.fromkeys(job.manual_questions.splitlines() if job.manual_questions else []))
        manual_questions.extend(open_questions)
        checklist.append(
            {
                "apply_order": index,
                "job_id": job.id,
                "company": job.company,
                "title": job.title,
                "status": job.status,
                "apply_url": job.apply_url,
                "location": job.location,
                "work_model": job.work_model,
                "fit_score": job.fit_score,
                "grade": grade.overall_grade if grade else "",
                "ready_to_send": grade.ready_to_send if grade else False,
                "materials": job.materials,
                "manual_questions": "\n".join(dict.fromkeys(manual_questions)),
                "resume_pdf_path": artifact.resume_pdf_path,
                "cover_letter_path": artifact.cover_letter_path,
                "resume_tex_path": artifact.resume_tex_path,
                "posting_age_text": job.posting_age_text,
                "posted_at": job.posted_at.isoformat() if job.posted_at else "",
                "first_seen_at": job.first_seen_at.isoformat(),
            }
        )
    return checklist


def export_current_json_csv(db: Session) -> dict[str, Path]:
    rows = current_export_rows(db)
    JSON_EXPORT_PATH.write_text(json.dumps(rows, indent=2), encoding="utf-8")
    if rows:
        with CSV_EXPORT_PATH.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
    else:
        CSV_EXPORT_PATH.write_text("", encoding="utf-8")
    return {"json": JSON_EXPORT_PATH, "csv": CSV_EXPORT_PATH}
